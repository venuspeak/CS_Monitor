# -*- coding : utf-8 -*-

import datetime
import json
import os
import time
import requests
import socket
import random
import copy
from openpyxl import Workbook, load_workbook

global OK
global zz
global all_attackIP_copy  # 复制一份 all_attackIP 才能进行删除，否则会报错
global today_attackIp_copy  # 复制一份 today_attackIP
dev_list = ['ip:session']  # cs设备的IP跟session
second = 120  # 请求获取CS日志的推迟时间，second、minute、hour
minute = 0
hour = 0
req_num = 20000  # 单次日志请求的条数最大为20000条
timeout = 100  # http请求响应超时时间
today_attackIp = set()  # 定义一个set集合，存储今天所有的攻击
all_attackIP = set()  # 定义一个set集合，存储所有的攻击IP
remove_attackIP = set()  # 定义个set集合，存储移除掉的IP
http_proxy = '代理地址'  # 获取IP归属外网代理，只发送告警事件到飞书群。否则，代理也发送的误报的话，日志太多，代理会崩溃的
req_header = {"Content-Type": "application/json;charset=UTF-8",
              "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:100.0) Gecko/20100101 Firefox/100.0"}  # 请求头时MIME是 application/json
interval = 120  # 间隔时间
allLogNumbers = []  # 存储每次监控频率处理的日志条数
timeList = ['00:00:00', '01:00:00', '02:00:00', '03:00:00', '04:00:00', '05:00:00', '06:00:00', '07:00:00',
            '08:00:00', '09:00:00', '10:00:00', '11:00:00', '12:00:00', '13:00:00', '14:00:00', '15:00:00',
            '16:00:00', '17:00:00', '18:00:00', '19:00:00', '20:00:00', '21:00:00', '22:00:00', '23:00:00',
            '23:59:30', '23:40:00']

sourIP_whiteList = ['']  # 研判分析表中，源IP为
destIP_whiteList = ['']  # 研判分析表中，目的IP
dev_whiteList = ['']  # 研判分析表中，源IP为安全设备转发(无需重点关注)
proxy_whiteList = ['']  # 研判分析表中，IP为公网代理地址(无需重点关注)
overload_whiteList = ['']  # 研判分析表中，IP为负载域的地址（判断规则是，将这些源IP设置为白名单,目的IP不设置)
WhiteList1 = []  # 白名单列表1
WhiteList2 = []  # 白名单列表2
'''
sour_ip not in sourIP_whiteList 
sour_ip not in dev_whiteList 
sour_ip not in proxy_whiteList 
sour_ip not in overload_whiteList
dest_ip not in destIP_whiteList 
dest_ip not in proxy_whiteList 
dest_ip not in sourIP_whiteList
'''
for abc in sourIP_whiteList:
    WhiteList1.append(abc)
for abc in dev_whiteList:
    WhiteList1.append(abc)
for abc in proxy_whiteList:
    WhiteList1.append(abc)
for abc in overload_whiteList:
    WhiteList1.append(abc)

for xyz in sourIP_whiteList:
    WhiteList2.append(xyz)
for xyz in destIP_whiteList:
    WhiteList2.append(xyz)
for xyz in proxy_whiteList:
    WhiteList2.append(xyz)


def monitor():
    # 开始对CS设备的日志进行数据请求

    req_url = "http://" + dev_ip + "/api/netidsEventlog/page"
    req_data = '{"datasourceType":"old","oldDatasourceType":"self","displayColumns":{"eventtypeid":1,"eventlevel":1,"eventtime":1,"attackid":1,"srcipStr":1,"dstipStr":1,"attackResult":1},"queryParam":{"startTime":"' + req_startTime + '","endTime":"' + req_endTime + '"},"page":{"pageNo":1,"pageSize":' + str(
        req_num) + '},"orderBy":{"field":"EVENTTIME","order":-1}}'
    get_log = conn.post(url=req_url, data=req_data, cookies=req_cookie, headers=req_header,
                        timeout=timeout).content.decode('utf-8', 'ignore')
    json_data = json.loads(get_log)

    # 开始对得到json数据进行处理
    event_dict = json_data['data']['records']  # 主要处理CS日志中 data中的records 字段数据

    allLogNumbers.append(len(event_dict))  # 将每次监控频率的条数存储

    num_url = 'http://' + dev_ip + '/api/netidsEventlog/analysis'
    num_data = '{"datasourceType":"old","oldDatasourceType":"day","queryParam":{}}'
    get_numbers = conn.post(url=num_url, data=num_data, cookies=req_cookie, timeout=timeout,
                            headers=req_header).content.decode('utf-8', 'ignore')
    num1 = json.loads(get_numbers)
    l1 = num1['data']['low']
    m1 = num1['data']['mid']
    h1 = num1['data']['high']
    day_num = l1 + m1 + h1

    global OK   # 今日日志完成率 = 脚本处理的日志条数 / cs上的日志总数
    if day_num > 0:  # day_num 必须大于0 ，否则会报错
        OK = '{:.2%}'.format(sum(allLogNumbers) / day_num)
        print('今日累计的日志条数: ' + str(day_num) + '条')
    else:
        OK = '当前时间段还没有事件'
    if str(datetime.datetime.now())[11:19] == '00:00:00':
        allLogNumbers.clear()

    eventDicts = reversed(event_dict)  # 得到的CS日志数据中时间是倒序，所以要反转列表（更改请求url中的order字段为1也可以），列表中的每个元素是每个单独的日志事件
    if len(event_dict) > 0:  # 如果日志条数大于0 ，则交给函数 managerData 处理
        out = "S/" + req_startTime + "     E/" + req_endTime + "   共" + str(len(event_dict)) + "条数据，正在处理..."
        out1 = '今日累计处理条数' + str(sum(allLogNumbers)) + '条，有效日志完成率为' + OK
        with open(filePath1, 'a', encoding='utf-8') as f1:  # 记录屏幕信息
            f1.write('\n' + out + '\n' + out1 + '\n')
        print(out)
        print(out1)
        managerData(eventDicts)
    else:
        out = "S/" + req_startTime + "     E/" + req_endTime + "   空数据"
        with open(filePath1, 'a', encoding='utf-8') as f1:  # 记录屏幕信息
            f1.write(out + '\n')
        print(out)


# 根据事件等级划分
def managerData(eventDict):
    count = 0  # 已合并IP

    lowLevel_sourList = []  # 将等级为 低级 的事件源IP 放入列表
    lowLevel_destList = []  # 将等级为 低级 的事件目的IP 放入列表
    lowLevel_list = []  # 将源、目的Ip 列表放入该列表
    midLevel_sourList = []
    midLevel_destList = []
    midLeve_list = []
    highLevel_sourList = []
    highLevel_destList = []
    highLevel_list = []
    # level_list = []
    for var in eventDict:
        event_level = var['eventlevel']
        sour_ip = var['srcipStr']
        dest_ip = var['dstipStr']
        if sour_ip not in WhiteList1:
            if dest_ip not in WhiteList2:
                if sour_ip != '已合并':
                    today_attackIp.add(sour_ip)
                    all_attackIP.add(sour_ip)
                    # if event_level > 0:
                    #     level_list.append(sour_ip)

                    if event_level == 20:
                        lowLevel_sourList.append(sour_ip)
                        lowLevel_destList.append(dest_ip)
                    elif event_level == 30:
                        midLevel_sourList.append(sour_ip)
                        midLevel_destList.append(dest_ip)
                    elif event_level == 40:
                        highLevel_sourList.append(sour_ip)
                        highLevel_destList.append(dest_ip)
                    elif event_level == 10:
                        pass
                    else:
                        pass
                else:
                    count += 1
                    out = '已合并IP共计 ' + str(count) + ' 个'
                    with open(filePath1, 'a', encoding='utf-8') as f1:  # 记录屏幕信息
                        f1.write(out + '\n')
                    print(out)
            else:
                pass
        else:
            pass
    lowLevel_list.append(lowLevel_sourList)
    lowLevel_list.append(lowLevel_destList)
    midLeve_list.append(midLevel_sourList)
    midLeve_list.append(midLevel_destList)
    highLevel_list.append(highLevel_sourList)
    highLevel_list.append(highLevel_destList)
    judge_lowLevel(lowLevel_list)
    judge_middleLevel(midLeve_list)
    judge_highLevel(highLevel_list)

    out123 = '\n' + '截止' + req_endTime + '攻击的IP一共有' + str(len(today_attackIp)) + '个   分别是:' + str(today_attackIp)
    out456 = '截止' + req_endTime + '移除的攻击IP一共有' + str(len(remove_attackIP)) + '个   分别是:' + str(remove_attackIP)
    with open(filePath1, 'a', encoding='utf-8') as f1:
        f1.write(out123 + '\n' + out456 + '\n\n')
    print(out123)
    print('')
    print(out456)


def judge_lowLevel(lowLevel):
    sourIP = list(set(lowLevel[0]))  # 获取lowLevel列表中的第一个元素(即源IP列表)
    sourIP = sorted(sourIP, key=socket.inet_aton)  # 对 ip 地址排序
    a = str(sourIP)
    b = a.replace("'", "\"")  # 替换掉set列表中的单引号为双引号
    c = b.replace(" ", "")  # 去掉set列表中的空格
    # destIP = set(midLevel[1])
    queryAll_url = "http://" + dev_ip + "/api/netidsEventlog/analysis"
    queryAll_data = '{"datasourceType":"old","oldDatasourceType":"day","queryParam":{"eventlevels":[20],"srcip6s":' + c + ',"startTime":"' + req_startTime + '","endTime":"' + req_endTime + '"}}'
    log_data = conn.post(url=queryAll_url, data=queryAll_data, cookies=req_cookie, timeout=timeout,
                         headers=req_header).content.decode('utf-8', 'ignore')
    allData = json.loads(log_data)['data']['low']
    if allData > 0:
        out1 = "当前时间段共计低危 " + str(allData) + " 条，详细如下:"
        with open(filePath2, 'a', encoding='utf-8') as f:
            f.write(out1 + '\n')
        for var in sourIP:
            query_data = '{"datasourceType":"old","oldDatasourceType":"day","queryParam":{"eventlevels":[20],"srcip6s":["' + var + '"],"startTime":"' + req_startTime + '","endTime":"' + req_endTime + '"}}}}'
            single_ip = conn.post(url=queryAll_url, data=query_data, cookies=req_cookie, timeout=timeout,
                                  headers=req_header).content.decode('utf-8', 'ignore')
            data = json.loads(single_ip)['data']['low']
            # print(data)
            out2 = "IP: " + var + "        "
            out2 = out2[:24] + "低危: " + str(data) + " 条"
            with open(filePath2, 'a', encoding='utf-8') as f:
                f.write(out2 + '\n')
    else:
        out1 = "当前时间段共计低危 0 条"
        with open(filePath2, 'a', encoding='utf-8') as f:
            f.write(out1 + '\n')


def judge_middleLevel(midLevel):
    sourIP = list(set(midLevel[0]))  # 获取midLevel列表中的第一个元素(即源IP列表)
    sourIP = sorted(sourIP, key=socket.inet_aton)  # 对 ip 地址排序
    a = str(sourIP)
    b = a.replace("'", "\"")  # 替换掉set列表中的单引号为双引号
    c = b.replace(" ", "")  # 去掉set列表中的空格
    # destIP = set(midLevel[1])
    queryAll_url = "http://" + dev_ip + "/api/netidsEventlog/analysis"
    queryAll_data = '{"datasourceType":"old","oldDatasourceType":"day","queryParam":{"eventlevels":[30],"srcip6s":' + c + ',"startTime":"' + req_startTime + '","endTime":"' + req_endTime + '"}}'
    log_data = conn.post(url=queryAll_url, data=queryAll_data, cookies=req_cookie, timeout=timeout,
                         headers=req_header).content.decode('utf-8', 'ignore')
    allData = json.loads(log_data)['data']['mid']
    if allData > 0:
        out1 = "当前时间段共计中危 " + str(allData) + " 条，详细如下:"
        with open(filePath2, 'a', encoding='utf-8') as f:
            f.write(out1 + '\n')
        for var in sourIP:
            query_data = '{"datasourceType":"old","oldDatasourceType":"day","queryParam":{"eventlevels":[30],"srcip6s":["' + var + '"],"startTime":"' + req_startTime + '","endTime":"' + req_endTime + '"}}}}'
            single_ip = conn.post(url=queryAll_url, data=query_data, cookies=req_cookie, timeout=timeout,
                                  headers=req_header).content.decode('utf-8', 'ignore')
            data = json.loads(single_ip)['data']['mid']
            out2 = "IP: " + var + "        "
            out2 = out2[:24] + "中危: " + str(data) + " 条"
            with open(filePath2, 'a', encoding='utf-8') as f:
                f.write(out2 + '\n')
    else:
        out1 = "当前时间段共计中危 0 条"
        with open(filePath2, 'a', encoding='utf-8') as f:
            f.write(out1 + '\n')


def judge_highLevel(highLevel):
    sourIP = list(set(highLevel[0]))  # 获取highLevel列表中的第一个元素(即源IP列表)
    sourIP = sorted(sourIP, key=socket.inet_aton)  # 对 ip 地址排序
    a = str(sourIP)
    b = a.replace("'", "\"")  # 替换掉set列表中的单引号为双引号
    c = b.replace(" ", "")  # 去掉set列表中的空格
    # destIP = set(midLevel[1])
    queryAll_url = "http://" + dev_ip + "/api/netidsEventlog/analysis"
    queryAll_data = '{"datasourceType":"old","oldDatasourceType":"day","queryParam":{"eventlevels":[30],"srcip6s":' + c + ',"startTime":"' + req_startTime + '","endTime":"' + req_endTime + '"}}'
    log_data = conn.post(url=queryAll_url, data=queryAll_data, cookies=req_cookie, timeout=timeout,
                         headers=req_header).content.decode('utf-8', 'ignore')
    allData = json.loads(log_data)['data']['high']
    if allData > 0:
        out1 = "当前时间段共计中危 " + str(allData) + " 条，详细如下:"
        with open(filePath2, 'a', encoding='utf-8') as f:
            f.write(out1 + '\n')
        for var in sourIP:
            query_data = '{"datasourceType":"old","oldDatasourceType":"day","queryParam":{"eventlevels":[30],"srcip6s":["' + var + '"],"startTime":"' + req_startTime + '","endTime":"' + req_endTime + '"}}}}'
            single_ip = conn.post(url=queryAll_url, data=query_data, cookies=req_cookie, timeout=timeout,
                                  headers=req_header).content.decode('utf-8', 'ignore')
            data = json.loads(single_ip)['data']['high']
            out2 = "IP: " + var + "        "
            out2 = out2[:24] + "高危: " + str(data) + " 条"
            with open(filePath2, 'a', encoding='utf-8') as f:
                f.write(out2 + '\n')
    else:
        out1 = "当前时间段共计高危 0 条"
        with open(filePath2, 'a', encoding='utf-8') as f:
            f.write(out1 + '\n')


def analysisEvent(rTime):
    x_x = sendTime(rTime)
    if x_x[0] != x_x[1]:
        global today_attackIp_copy, all_attackIP_copy
        today_attackIp_copy = copy.deepcopy(today_attackIp)
        all_attackIP_copy = copy.deepcopy(all_attackIP)
        for var in today_attackIp:
            sourIPList = set()  # 创建源ip列表，查询一个攻击ip 对应多少个攻击ip
            destIPList = set()  # 创建目的ip列表，查询一个攻击ip 对应多少个攻击ip
            eventTimeList = []  # 攻击时间列表
            eventIDList = []  # 攻击事件ID列表
            eventNameList = set()  # 攻击事件类型列表
            lowLevelList = set()
            midLevelList = set()
            highLevelList = set()

            # 查询当天的日志条数及具体事件,一个IP处理前10000条数据
            queryLog_url = 'http://' + dev_ip + '/api/netidsEventlog/page'
            queryLog_day = '{"datasourceType":"old","oldDatasourceType":"day","displayColumns":{"srcport":1,"dstport":1,"eventtypeid":1,"eventlevel":1,"eventtime":1,"securityid":1,"attackid":1,"srcipStr":1,"dstipStr":1,"attackResult":1},"queryParam":{"srcip6s":["' + var + '"]},"page":{"pageNo":1,"pageSize":10000},"orderBy":{"field":"EVENTTIME","order":-1}}'

            queryLog1 = conn.post(url=queryLog_url, data=queryLog_day, cookies=req_cookie, timeout=timeout,
                                  headers=req_header).content.decode('utf-8', 'ignore')

            logDict11 = json.loads(queryLog1)['data']['records']
            logDict1 = reversed(logDict11)

            for var1 in logDict1:
                sour_ip = var1['srcipStr']  # 攻击IP
                sourIPList.add(sour_ip)
                dest_ip = var1['dstipStr']  # 受害IP
                destIPList.add(dest_ip)
                event_id = var1['recid']  # 事件id
                eventIDList.append(event_id)
                event_time = var1['eventTimeStr']  # 事件发生时间
                eventTimeList.append(event_time)
                event_name = var1['eventName']  # 事件名称
                eventNameList.add(event_name)
                event_level = var1['eventlevel']  # 攻击事件等级
                if event_level == 20:
                    lowLevelList.add(event_name)
                elif event_level == 30:
                    midLevelList.add(event_name)
                elif event_level == 40:
                    highLevelList.add(event_name)
                else:
                    pass
            if len(sourIPList) > 0:  # 因为使用了set列表中的remove方法，删除set表中的元素后，元素会占位，所以有空元素出现

                x = detailEvent(eventIDList, eventTimeList, sourIPList, destIPList, eventNameList, len(logDict11),
                                lowLevelList,
                                midLevelList, highLevelList)
                # print('开始发送数据到表格')
                xlsxOutContent(x, rTime)
    else:
        # print('ip清零之前' + str(today_attackIp))
        with open(datetime.datetime.now().strftime('%Y_%m') + '_today.log', 'a', encoding='utf-8') as attack:
            i = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            v = today_attackIp
            k = all_attackIP
            out = '截止: ' + str(i) + '\n' + '今日攻击IP共 ' + str(len(today_attackIp)) + ' 个已被清零，具体IP如下:   ' + str(v)
            out1 = '从脚本开始运行时到现在，已发现攻击IP累计 ' + str(len(all_attackIP)) + '个，具体如下:   ' + str(k)
            attack.write(out + '\n\n' + out1 + '\n\n')
        today_attackIp.clear()
        remove_attackIP.clear()

        # print('ip清零之后' + str(today_attackIp))
        print('今日攻击IP列表已清零')


# 判断事件的具体细节字段
def detailEvent(eventIDList, eventTimeList, sourIPList, destIPList, eventNameList, logDict, lowLevelList, midLevelList,
                highLevelList):
    portList1 = []
    portList2 = []

    # 扫描事件的列表
    scanList = ['SCAN_SYNONLY_TCP端口扫描', 'SCAN_UDP端口扫描', 'SCAN_ICMP扫描探测']
    xsList = ['HTTP_SQL注入攻击', 'HTTP_XSS疑似注入', 'HTTP_url躲避']
    if len(eventNameList) == 1 and list(eventNameList)[0] in scanList:
        # 只处理前10条的事件详情
        j = eventIDList[:10]
        k = eventTimeList[:10]
    elif len(eventNameList) == 1 and list(eventNameList)[0] in xsList:
        # 只处理前20条的事件详情
        j = eventIDList[:20]
        k = eventTimeList[:20]
    else:
        # 只处理前30条的事件详情
        j = eventIDList[:30]
        k = eventTimeList[:30]

    # 事件细节内容,get请求
    for m, n in zip(j, k):
        detail_url = 'http://' + dev_ip + '/api/netidsEventlog/detail/old/' + m + '/' + n
        queryDetail = conn.get(url=detail_url, cookies=req_cookie, timeout=timeout,
                               headers=req_header).content.decode('utf-8', 'ignore')
        detailDict = json.loads(queryDetail)['data']

        port1 = detailDict['srcport']  # 攻击IP端口
        port2 = detailDict['dstport']  # 受害IP端口
        portList1.append(port1)
        portList2.append(port2)

    # 规则开始定义 规则开始定义 规则开始定义 规则开始定义 规则开始定义 规则开始定义 规则开始定义
    # 提示：可以根事件列表、源IP、目的IP、源端口、目的端口定义
    # 受害IP为1个，源IP端口只有一个，事件名称只有一个，则判定为误报
    if len(portList1) <= 2 and 1234 in list(set(portList1)):
        returnList2 = set()
        for mmm, nnn in zip(j, k):
            detail_url = 'http://' + dev_ip + '/api/netidsEventlog/detail/old/' + mmm + '/' + nnn
            queryDetail = conn.get(url=detail_url, cookies=req_cookie, timeout=timeout,
                                   headers=req_header).content.decode('utf-8', 'ignore')
            detailDict1 = json.loads(queryDetail)['data']
            returnInfo1 = detailDict1['eventlevel']  # 攻击IP 返回信息
            returnList2.add(returnInfo1)

        # 如果返回的returnLevel 等级都为 40 的话，则判断为误报
        if len(destIPList) == 1 and len(sourIPList) == 1 and len(returnList2) == 1 and list(returnList2)[0] == 40:
            v1 = '截止到今天' + datetime.datetime.now().strftime('%H:%M:%S') + '源IP: ' + list(sourIPList)[0] + ' 目的IP: ' + \
                 list(destIPList)[0] + '\n' + '攻击端口固定为: ' + str(
                list(set(portList1))[
                    0]) + '\n' + '攻击事件名称为: ' + list(eventNameList)[0] + '\n' + '这个事件大概率是CS误报，请及时核查！！！！！！！！' + '\n'
            sendSingle(v1)
            # print(v1)
            today_attackIp_copy.discard(list(sourIPList)[0])
            all_attackIP_copy.discard(list(sourIPList)[0])
            remove_attackIP.add(list(sourIPList)[0])

    x = list(set(portList1))
    y = list(set(portList2))
    p = list(sourIPList)[0]
    q = list(destIPList)

    # print('一次传送数据'+str(x), str(y), p, q, lowLevelList, midLevelList, highLevelList)
    return x, y, p, q, lowLevelList, midLevelList, highLevelList


def xlsxOutContent(x, rTime):  # 将告警事件记录到表格跟发送给飞书
    a1 = rTime[0:4]  # 2022 年
    b1 = rTime[5:7]  # 05 月
    c1 = rTime[8:10]  # 20 日
    d1 = rTime[11:13]  # 23 点
    e1 = rTime[14:16]  # 59 分
    f1 = rTime[17:19]  # 59 秒
    g1 = datetime.date(int(a1), int(b1), int(c1))
    h1 = datetime.time(int(d1), int(e1), int(f1))
    sTime = datetime.datetime.combine(g1, h1).strftime('%Y-%m-%d %H:%M:%S')  # 重新写时间

    # 两个月
    eTime_m2 = (datetime.datetime.combine(g1, h1) - datetime.timedelta(days=60)).strftime('%Y-%m-%d %H:%M:%S')
    # 三个月
    eTime_m3 = (datetime.datetime.combine(g1, h1) - datetime.timedelta(days=90)).strftime('%Y-%m-%d %H:%M:%S')
    # 半年
    eTime_m6 = (datetime.datetime.combine(g1, h1) - datetime.timedelta(days=180)).strftime('%Y-%m-%d %H:%M:%S')
    # 一年
    eTime_m12 = (datetime.datetime.combine(g1, h1) - datetime.timedelta(days=365)).strftime('%Y-%m-%d %H:%M:%S')

    if x[2] in today_attackIp_copy:
        print('表格中已记录了数据1条')
        xlsxOutList = [dev_ip, str(x[2]), str(x[3]), str(x[0]), str(x[1])]

        queryLog_url = 'http://' + dev_ip + '/api/netidsEventlog/page'
        queryLog_year = '{"datasourceType":"old","oldDatasourceType":"self","displayColumns":{"srcport":1,"dstport":1,"eventtypeid":1,"eventlevel":1,"eventtime":1,"securityid":1,"attackid":1,"srcipStr":1,"dstipStr":1,"attackResult":1},"queryParam":{"srcip6s":["' + \
                        x[
                            2] + '"],"startTime":"' + eTime_m12 + '","endTime":"' + sTime + '"},"page":{"pageNo":1,"pageSize":10},"orderBy":{"field":"EVENTTIME","order":1}}'
        queryLog2 = conn.post(url=queryLog_url, data=queryLog_year, cookies=req_cookie, timeout=timeout,
                              headers=req_header).content.decode('utf-8', 'ignore')
        logDict2 = json.loads(queryLog2)['data']['records']
        event_time1 = logDict2[0]['eventTimeStr']
        xlsxOutList.append(event_time1)  # 一年之内最早的一条事件攻击时间

        # 查询分析低中高危事件的url 跟 请求数据
        query_url = "http://" + dev_ip + "/api/netidsEventlog/analysis"
        query_day = '{"datasourceType":"old","oldDatasourceType":"day","queryParam":{"srcip6s":["' + x[2] + '"]}}'
        query_week = '{"datasourceType":"old","oldDatasourceType":"week","queryParam":{"srcip6s":["' + x[2] + '"]}}'
        query_month = '{"datasourceType":"old","oldDatasourceType":"month","queryParam":{"srcip6s":["' + x[2] + '"]}}'
        query_m2 = '{"datasourceType":"old","oldDatasourceType":"self","queryParam":{"srcip6s":["' + x[
            2] + '"],"startTime":"' + eTime_m2 + '","endTime":"' + sTime + '"}}'
        query_m3 = '{"datasourceType":"old","oldDatasourceType":"self","queryParam":{"srcip6s":["' + x[
            2] + '"],"startTime":"' + eTime_m3 + '","endTime":"' + sTime + '"}}'
        query_m6 = '{"datasourceType":"old","oldDatasourceType":"self","queryParam":{"srcip6s":["' + x[
            2] + '"],"startTime":"' + eTime_m6 + '","endTime":"' + sTime + '"}}'
        query_m12 = '{"datasourceType":"old","oldDatasourceType":"self","queryParam":{"srcip6s":["' + x[
            2] + '"],"startTime":"' + eTime_m12 + '","endTime":"' + sTime + '"}} '

        # 当天
        query1 = conn.post(url=query_url, data=query_day, cookies=req_cookie, timeout=timeout,
                           headers=req_header).content.decode('utf-8', 'ignore')
        num1 = json.loads(query1)
        low1 = num1['data']['low']
        mid1 = num1['data']['mid']
        high1 = num1['data']['high']
        day_num = str(low1 + mid1 + high1)
        xlsxOutList.append(day_num)  # 当天攻击的总事件条数

        if len(x[4]) != 0:
            xlsxOutList.append(str(list(x[4])))  # 当天低危攻击事件类型
        else:
            xlsxOutList.append('')
        if len(x[5]) != 0:
            xlsxOutList.append(str(list(x[5])))  # 当天中危攻击事件类型
        else:
            xlsxOutList.append('')
        if len(x[6]) != 0:
            xlsxOutList.append(str(list(x[6])))  # 当天高危攻击事件类型
        else:
            xlsxOutList.append('')

        # 本周
        query2 = conn.post(url=query_url, data=query_week, cookies=req_cookie, timeout=timeout,
                           headers=req_header).content.decode('utf-8', 'ignore')
        num2 = json.loads(query2)
        low2 = num2['data']['low']
        mid2 = num2['data']['mid']
        high2 = num2['data']['high']
        week_num = str(low2 + mid2 + high2)
        xlsxOutList.append(week_num)  # 本周攻击的总事件条数

        # 本月
        query3 = conn.post(url=query_url, data=query_month, cookies=req_cookie, timeout=timeout,
                           headers=req_header).content.decode('utf-8', 'ignore')
        num3 = json.loads(query3)
        low3 = num3['data']['low']
        mid3 = num3['data']['mid']
        high3 = num3['data']['high']
        m1_num = str(low3 + mid3 + high3)
        xlsxOutList.append(m1_num)  # 本月攻击的总事件条数

        # 两个月
        query4 = conn.post(url=query_url, data=query_m2, cookies=req_cookie, timeout=timeout,
                           headers=req_header).content.decode('utf-8', 'ignore')
        num4 = json.loads(query4)
        low4 = num4['data']['low']
        mid4 = num4['data']['mid']
        high4 = num4['data']['high']
        m2_num = str(low4 + mid4 + high4)
        xlsxOutList.append(m2_num)  # 两个月攻击的总事件条数

        # 三个月
        query5 = conn.post(url=query_url, data=query_m3, cookies=req_cookie, timeout=timeout,
                           headers=req_header).content.decode('utf-8', 'ignore')
        num5 = json.loads(query5)
        low5 = num5['data']['low']
        mid5 = num5['data']['mid']
        high5 = num5['data']['high']
        m3_num = str(low5 + mid5 + high5)
        xlsxOutList.append(m3_num)  # 三个月攻击的总事件条数

        # 六个月
        query6 = conn.post(url=query_url, data=query_m6, cookies=req_cookie, timeout=timeout,
                           headers=req_header).content.decode('utf-8', 'ignore')
        num6 = json.loads(query6)
        low6 = num6['data']['low']
        mid6 = num6['data']['mid']
        high6 = num6['data']['high']
        m6_num = str(low6 + mid6 + high6)
        xlsxOutList.append(m6_num)  # 六个月攻击的总事件条数

        # 12个月
        query7 = conn.post(url=query_url, data=query_m12, cookies=req_cookie, timeout=timeout,
                           headers=req_header).content.decode('utf-8', 'ignore')
        num7 = json.loads(query7)
        low7 = num7['data']['low']
        mid7 = num7['data']['mid']
        high7 = num7['data']['high']
        m12_num = str(low7 + mid7 + high7)
        xlsxOutList.append(m12_num)  # 一年之内攻击的事件条数

        # 发送事件列表xlsx
        xlsxOut(xlsxOutList)
        # print('已经写入了1行数据')

    x_x = sendTime(rTime)

    page_url = 'http://' + dev_ip + '/api/netidsEventlog/page'
    analysis_url = 'http://' + dev_ip + '/api/netidsEventlog/analysis'
    # print('\n' + '开始发送告警事件到飞书。 其中今日累计攻击有' + str(len(today_attackIp)) + '条。移除掉误报的IP后有' + str(len(today_attackIp_copy)) + '条') 冲突了，today_attackIp_copy在修改中不能执行
    if x[2] in today_attackIp_copy:
        print('开始向飞书发送告警')
        outEvent = []  # 输出的飞书数据
        out10 = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        outEvent.append(out10)
        page_data1 = '{"datasourceType":"old","oldDatasourceType":"self","displayColumns":{"srcport":1,"dstport":1,"eventtypeid":1,"eventlevel":1,"eventtime":1,"securityid":1,"attackid":1,"srcipStr":1,"dstipStr":1,"attackResult":1},"queryParam":{"srcip6s":["' + \
                     x[2] + '"],"startTime":"' + x_x[0] + '","endTime":"' + x_x[
                         1] + '"},"page":{"pageNo":1,"pageSize":' + str(
            req_num) + '},"orderBy":{"field":"EVENTTIME","order":-1}}'
        analysis_data1 = '{"datasourceType":"old","oldDatasourceType":"self","queryParam":{"srcip6s":["' + x[
            2] + '"],"startTime":"' + \
                         x_x[0] + '","endTime":"' + x_x[1] + '"}}'
        # str2 = '{"datasourceType":"old","oldDatasourceType":"self","queryParam":{"srcip6s":["' + var + '"],"startTime":"'+req_startTime[:10]+' 00:00:00","endTime":"'+req_endTime[:10]+' 03:00:00"}}'
        # str1 = '{"datasourceType":"old","oldDatasourceType":"self","displayColumns":{"srcport":1,"dstport":1,"eventtypeid":1,"eventlevel":1,"eventtime":1,"securityid":1,"attackid":1,"srcipStr":1,"dstipStr":1,"attackResult":1},"queryParam":{"srcip6s":["'+var+'"],"startTime":"'+req_startTime[:10]+' 00:00:00","endTime":"'+req_endTime[:10]+' 03:00:00"},"page":{"pageNo":1,"pageSize":100},"orderBy":{"field":"EVENTTIME","order":-1}}'

        analysis_log = conn.post(url=analysis_url, data=analysis_data1, cookies=req_cookie, headers=req_header,
                                 timeout=timeout).content.decode('utf-8', 'ignore')
        # print(analysis_log)
        high_num = json.loads(analysis_log)['data']['high']
        mid_num = json.loads(analysis_log)['data']['mid']
        low_num = json.loads(analysis_log)['data']['low']
        out5 = '源ip: ' + x[2] + ' 累计事件' + str(high_num + mid_num + low_num) + '条,其中'
        outEvent.append(out5)

        if high_num + mid_num + low_num >= 8000:
            outEvent.append('***每次检索8000条数据***')

        page_log = conn.post(url=page_url, data=page_data1, cookies=req_cookie, headers=req_header,
                             timeout=timeout).content.decode('utf-8', 'ignore')
        eventDict = json.loads(page_log)['data']['records']
        eventDict = reversed(eventDict)

        high_eventName = []
        mid_eventName = []
        low_eventName = []
        sourIP_list = []
        destIP_list = []

        for var1 in eventDict:

            event_name = var1['eventName']
            dest_ip = var1['dstipStr']
            event_level = var1['eventlevel']
            destIP_list.append(dest_ip)
            if event_level == 20:
                low_eventName.append(event_name)
            elif event_level == 30:
                mid_eventName.append(event_name)
            elif event_level == 40:
                high_eventName.append(event_name)
            else:
                pass

        # print("攻击IP: " + var + "  "+"受害IP:" + str(list(set(destIP_list))))
        if len(list(set(low_eventName))) != 0:
            out1 = '低危: {} 条  事件: {}'.format(str(len(low_eventName)), str(list(set(low_eventName))))
            outEvent.append(out1)
        else:
            pass
        if len(list(set(mid_eventName))) != 0:
            out2 = '中危: {} 条  事件: {}'.format(str(len(mid_eventName)), str(list(set(mid_eventName))))
            outEvent.append(out2)
        else:
            pass
        if len(list(set(high_eventName))) != 0:
            out3 = '高危: {} 条  事件: {}'.format(str(len(high_eventName)), str(list(set(high_eventName))))
            outEvent.append(out3)
        else:
            pass

        out6 = '\n' + '告警设备: ' + dev_ip
        outEvent.append(out6)
        out7 = '有效日志完成率: ' + OK
        outEvent.append(out7)
        # print("低危: " + str(len(low_eventName)) + "条     事件:" + str(list(set(low_eventName))))
        # print("中危: " + str(len(mid_eventName)) + "条     事件:" + str(list(set(mid_eventName))))
        # print("高危: " + str(len(high_eventName)) + "条     事件:" + str(list(set(high_eventName))))
        # out4 = '截止: '+x[1] + '攻击的IP一共有' + str(len(today_attackIp)) + '条'
        sendFeishu(outEvent)


# 将误报记录到本地，不在发送到飞书误报群
def sendSingle(out):

    with open(filePath4, 'a', encoding='utf-8') as f1234:  # 记录屏幕信息
        f1234.write(out + '\n')


# 发送的表格，每隔一个小时创建一个表格，每次分析，以最新的表格为准
def xlsxOut(outList):
    try:
        currentPath = os.getcwd()  # 查看当前工作目录（绝对路径）
        tmpFolder = datetime.datetime.now().strftime('%Y_%m_%d')  # 创建每天的文件夹
        targetFolder = currentPath + os.path.sep + 'eventXlsx' + os.path.sep + tmpFolder

        name = datetime.datetime.now().strftime('%H') + '点' + '.xlsx'  # 每天最多24个文件

        if not os.path.exists(targetFolder):
            os.makedirs(targetFolder)  # 创建目录
        else:
            pass

        os.chdir(targetFolder)  # 修改当前工作目录
        path = os.listdir()
        # print('修改前表格路径 '+os.getcwd())

        if name not in path:
            wb = Workbook()
            ws = wb.active
            ws.append(['设备IP', '攻击ip', '受害ip', '攻击端口', '受害端口', '一年之内最早的一条事件攻击时间', '当天攻击的总事件条数',
                       '当天低危攻击事件类型', '当天中危攻击事件类型', '当天高危攻击事件类型', '本周攻击的总事件条数', '本月攻击的总事件条数', '两个月攻击的总事件条数',
                       '三个月攻击的总事件条数',
                       '六个月攻击的总事件条数', '一年之内攻击的事件条数'])
            ws.append(outList)
            wb.save(name)
            os.chdir(currentPath)
            # print('表格输出，完成1次')
            # print('修改后文件路径'+os.getcwd()+'数据已经输出到表格中----1')
        else:
            wb = load_workbook(name)
            ws = wb.active
            ws.append(outList)
            wb.save(name)
            os.chdir(currentPath)
            # print('修改后文件路径'+os.getcwd()+'数据已经输出到表格中----2')
            # print('表格输出，完成2次')
    except Exception as e1:
        print('写入表格的时候请勿打开表格，为记录上的xlsx数据，在当前文件夹中的备份xlsx表格查看!!!')
        currentPath = os.getcwd()  # 查看当前工作目录（绝对路径）
        tmpFolder = datetime.datetime.now().strftime('%Y_%m_%d')  # 创建每天的文件夹
        targetFolder = currentPath + os.path.sep + 'eventXlsx' + os.path.sep + tmpFolder

        name = datetime.datetime.now().strftime('%H') + '点_backup' + '.xlsx'  # 每天24个文件

        if not os.path.exists(targetFolder):
            os.makedirs(targetFolder)  # 创建目录
        else:
            pass

        os.chdir(targetFolder)  # 修改当前工作目录
        path = os.listdir()
        if name not in path:
            wb = Workbook()
            ws = wb.active
            ws.append(['设备IP', '攻击ip', '受害ip', '攻击端口', '受害端口', '一年之内最早的一条事件攻击时间', '当天攻击的总事件条数',
                       '当天低危攻击事件类型', '当天中危攻击事件类型', '当天高危攻击事件类型', '本周攻击的总事件条数', '本月攻击的总事件条数', '两个月攻击的总事件条数',
                       '三个月攻击的总事件条数',
                       '六个月攻击的总事件条数', '一年之内攻击的事件条数'])
            ws.append(outList)
            wb.save(name)
            os.chdir(currentPath)
        else:
            wb = load_workbook(name)
            ws = wb.active
            ws.append(outList)
            wb.save(name)
            os.chdir(currentPath)


# 处理发送给飞书群的时间
def sendTime(rTime):
    timeStr = 0
    # 处理时间
    a1 = rTime[0:4]  # 2022 年
    b1 = rTime[5:7]  # 05 月
    c1 = rTime[8:10]  # 20 日
    d1 = rTime[11:13]  # 23 点
    e1 = rTime[14:16]  # 59 分
    f1 = rTime[17:19]  # 59 秒
    g1 = datetime.date(int(a1), int(b1), int(c1))
    h1 = datetime.time(00, 00, 00)
    start = datetime.datetime.combine(g1, h1).strftime('%Y-%m-%d %H:%M:%S')  # 重新写时间

    if d1 == '01' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 1
    elif d1 == '02' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 2
    elif d1 == '03' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 3
    elif d1 == '04' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 4
    elif d1 == '05' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 5
    elif d1 == '06' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 6
    elif d1 == '07' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 7
    elif d1 == '08' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 8
    elif d1 == '09' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 9
    elif d1 == '10' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 10
    elif d1 == '11' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 11
    elif d1 == '12' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 12
    elif d1 == '13' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 13
    elif d1 == '14' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 14
    elif d1 == '15' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 15
    elif d1 == '16' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 16
    elif d1 == '17' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 17
    elif d1 == '18' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 18
    elif d1 == '19' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 19
    elif d1 == '20' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 20
    elif d1 == '21' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 21
    elif d1 == '22' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 22
    elif d1 == '23' and e1 == '00' and f1 == '00':
        timeStr = timeStr + 23
    elif d1 == '23' and e1 == '40' and f1 == '00':
        timeStr = timeStr + 24
    elif d1 == '23' and e1 == '59' and f1 == '30':
        timeStr = 0  # 当 sTime 跟 eTime 时间相等时，清除一天之内的攻击IP
    else:
        pass

    sTime = start
    eTime = (datetime.datetime.combine(g1, h1) + datetime.timedelta(hours=timeStr)).strftime('%Y-%m-%d %H:%M:%S')

    # if d1 == '23' and e1 == '59' and f1 == '30':  # 查询当天的所有攻击
    #     eTime = (datetime.datetime.combine(g1, h1) + datetime.timedelta(hours=23, minutes=59, seconds=55)).strftime(
    #         '%Y-%m-%d %H:%M:%S')
    # out_event.append(eTime)
    return sTime, eTime


# 发送给飞书群（钉钉一分钟发送20条数据）
def sendFeishu(out):
    ss = ''
    if len(out) == 8:
        ss = '截止到: ' + out[0] + '\n' + out[1] + '\n' + out[2] + '\n' + out[3] + '\n' + out[4] + '\n' + out[5] + '\n' + \
             out[
                 6] + '\n' + out[7] + '\n'
    elif len(out) == 7:
        ss = '截止到: ' + out[0] + '\n' + out[1] + '\n' + out[2] + '\n' + out[3] + '\n' + out[4] + '\n' + out[5] + '\n' + \
             out[
                 6] + '\n'
    elif len(out) == 6:
        ss = '截止到: ' + out[0] + '\n' + out[1] + '\n' + out[2] + '\n' + out[3] + '\n' + out[4] + '\n' + out[5] + '\n'
    elif len(out) == 5:
        ss = '截止到: ' + out[0] + '\n' + out[1] + '\n' + out[2] + '\n' + out[3] + '\n' + out[4]
    elif len(out) == 1:
        ss = out
    else:
        pass

    with open('sendFieshu.log', 'a', encoding='utf-8') as fff:
        fff.write(ss + '\n')

    headers = {'Content-Type': 'application/json',
               'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:100.0) Gecko/20100101 Firefox/100.0'}  # 定义数据类型
    webhook = 'https://open.feishu.cn/open-apis/'  # 定义webhook，从飞书群机器人设置页面复制获得
    webhook1 = 'https://open.feishu.cn/open-apis/'
    webhook2 = 'https://open.feishu.cn/open-apis/'
    webhook3 = 'https://open.feishu.cn/open-apis/'
    webhook4 = 'https://open.feishu.cn/open-apis/'
    webhook5 = 'https://open.feishu.cn/open-apis/'
    webhook6 = 'https://open.feishu.cn/open-apis/'
    webhook7 = 'https://open.feishu.cn/open-apis/'
    webhook8 = 'https://open.feishu.cn/open-apis/'
    webhookList = [webhook, webhook1, webhook2, webhook3, webhook4, webhook5, webhook6, webhook7, webhook8]
    # 定义要发送的数据
    try:
        data = {
            "msg_type": "text",
            "content": {"text": ss}
        }

        requests.post(random.choice(webhookList), proxies={'http': http_proxy, 'https': http_proxy},
                      data=json.dumps(data),
                      headers=headers)  # 发送post请求
    except Exception as e:
        data = {
            "msg_type": "text",
            "content": {"text": e}
        }
        req = requests.post(webhook, proxies={'http': http_proxy, 'https': http_proxy}, data=json.dumps(data),
                            headers=headers)  # 发送post请求
        print(req.text)


#  测试向飞书群发送第一条消息
def oneMessage():
    headers = {'Content-Type': 'application/json',
               'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:100.0) Gecko/20100101 Firefox/100.0'}  # 定义数据类型

    webhook = 'https://open.feishu.cn/open-apis/'  # 定义webhook，从飞书群机器人设置页面复制获得
    ss = '截止到' + str(datetime.datetime.now()) + '内网CS脚本最终版开始运行，向飞书群发送第一条消息'
    data = {
        "msg_type": "text",
        "content": {"text": ss}
    }

    req = requests.post(webhook, proxies={'http': http_proxy, 'https': http_proxy}, data=json.dumps(data),
                        headers=headers)  # 发送post请求
    print(req.text)


def SS():
    SSS = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if SSS[11:19] in timeList:
        analysisEvent(SSS)  # 匹配规则的事件发送到飞书后（即误报的事件）在将被告警的事件（即非误报）写入表格，然后将告警事件发送到飞书
        # print('开始时间是' + x[0])
        # print('结束时间是' + x[1])
        # xlsxOut(SSS)
    else:
        pass


if __name__ == "__main__":

    oneMessage()  # 发送给飞书群的测试消息

    # 创建四个文件夹，分别记录屏幕输出信息、监控频率日志、表格分析日志、误报事件日志。文件夹名分别为 screenTxt、monitorLog、eventXlsx、errorLog
    curPath = os.getcwd()
    tmp1 = 'screenTxt'
    tmp2 = 'monitorLog'
    tmp3 = 'eventXlsx'
    tmp4 = 'errorLog'
    target1 = curPath + os.path.sep + tmp1
    target2 = curPath + os.path.sep + tmp2
    target3 = curPath + os.path.sep + tmp3
    target4 = curPath + os.path.sep + tmp4
    if not os.path.exists(target1):
        os.makedirs(target1)
    else:
        pass
    if not os.path.exists(target2):
        os.makedirs(target2)
    else:
        pass
    if not os.path.exists(target3):
        os.makedirs(target3)
    else:
        pass
    if not os.path.exists(target4):
        os.makedirs(target4)
    else:
        pass

    req_endTime = (datetime.datetime.now() - datetime.timedelta(seconds=300, minutes=minute,
                                                                hours=hour)).strftime(
        '%Y-%m-%d %H:%M:%S')  # 结束请求时间比请求开始时间推迟120秒

    req_startTime = (datetime.datetime.now() - datetime.timedelta(seconds=second + 300, minutes=minute,
                                                                  hours=hour)).strftime(
        '%Y-%m-%d %H:%M:%S')  # 开始请求时间比当前系统时间推迟5分钟

    while True:

        fileName1 = "screenInfo" + (datetime.datetime.now() - datetime.timedelta(seconds=300)).strftime(
            '%Y_%m_%d') + ".txt"
        fileName2 = "monitor" + (datetime.datetime.now() - datetime.timedelta(seconds=300)).strftime(
            '%Y_%m_%d') + ".log"
        fileName3 = "xlsx" + (datetime.datetime.now() - datetime.timedelta(seconds=300)).strftime('%Y_%m_%d') + ".xlsx"
        fileName4 = "errorLog" + (datetime.datetime.now() - datetime.timedelta(seconds=300)).strftime(
            '%Y_%m_%d') + ".log"

        filePath1 = target1 + os.path.sep + fileName1
        filePath2 = target2 + os.path.sep + fileName2
        filePath3 = target3 + os.path.sep + fileName3
        filePath4 = target4 + os.path.sep + fileName4

        timeStart = time.time()
        conn = requests.session()  # 为了保持会话的连接，用session方法

        with open(filePath2, 'a', encoding='utf-8') as f:
            f.write('\n' + "S/" + req_startTime + "     E/" + req_endTime + '\n')
        for dev in dev_list:
            dev_data = dev.split(':')
            dev_ip = dev_data[0]
            dev_session = dev_data[1]
            req_cookie = {"JSESSIONID": dev_session}
            monitor()
        timeEnd = time.time()
        timeInterval = interval - (timeEnd - timeStart)
        if timeInterval > 0:
            xx = time.time()
            for ii in range(int(timeInterval)):  # 这段代码非常之重要
                SS()
                time.sleep(1)
            yy = time.time()
            zz = yy - xx
            print("内网CS监控测试，请勿关闭 当前频率为" + str(interval) + "秒每次，" + str(int(zz)) + "秒后继续......")
            print('')
            # time.sleep(zz)
        else:
            print("检测到延时丢包" + str(int(timeInterval)) + "秒")

        # 定义请求时间
        req_endTime = (datetime.datetime.now() - datetime.timedelta(seconds=300, minutes=minute,
                                                                    hours=hour)).strftime(
            '%Y-%m-%d %H:%M:%S')  # 结束请求时间比请求开始时间推迟120秒

        req_startTime = (datetime.datetime.now() - datetime.timedelta(seconds=second + 300, minutes=minute,
                                                                      hours=hour)).strftime(
            '%Y-%m-%d %H:%M:%S')  # 开始请求时间比当前系统时间推迟5分钟
        if zz > 120:  # 这里设置的代表不会错过任何一条日志
            req_startTime = (datetime.datetime.now() - datetime.timedelta(seconds=300 + zz, minutes=minute,
                                                                          hours=hour)).strftime(
                '%Y-%m-%d %H:%M:%S')  # 开始请求时间提前
